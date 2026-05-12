import csv
import io
from decimal import Decimal, InvalidOperation
import unicodedata

from app.services.cadastro_service import CadastroService


COLUNAS_IMPORTACAO = {
    'numero cnj': 'numero_processo',
    'numero do cnj': 'numero_processo',
    'n cnj': 'numero_processo',
    'n cnj': 'numero_processo',
    'numero processo': 'numero_processo',
    'numero processo judicial': 'numero_processo',
    'numero alternativo': 'numero_processo',
    'cnj': 'numero_processo',
    'processo': 'numero_processo',
    'pasta': 'titulo',
    'titulo': 'titulo',
    'descricao': 'descricao',
    'objetos': 'descricao',
    'assunto': 'descricao',
    'tipo de acao': 'tipo_acao',
    'classe': 'tipo_acao',
    'classe judicial': 'tipo_acao',
    'classe processual': 'tipo_acao',
    'natureza classe processual': 'tipo_acao',
    'natureza processual': 'tipo_acao',
    'tribunal': 'tribunal',
    'orgao judicial': 'tribunal',
    'origem': 'tribunal',
    'vara': 'vara',
    'escrivania secao': 'vara',
    'comarca': 'comarca',
    'valor da causa': 'valor_da_causa',
    'status': 'status',
    'fase': 'fase',
    'data de criacao': 'data_criacao',
    'data distribuicao': 'data_criacao',
    'distribuicao': 'data_criacao',
    'parte cliente': 'parte_cliente',
    'tipo da parte cliente': 'tipo_parte_cliente',
    'contato da parte cliente': 'contato_parte_cliente',
    'cpf cnpj parte cliente': 'contato_parte_cliente',
    'parte adversa': 'parte_adversa',
    'tipo da parte adversa': 'tipo_parte_adversa',
    'contato da parte adversa': 'contato_parte_adversa',
    'cpf cnpj parte adversa': 'contato_parte_adversa',
    'outras partes': 'outras_partes',
    'prazos': 'prazos',
    'intimacao': 'prazos',
    'defesa': 'prazos',
    'compromissos': 'prazos',
    'tarefas': 'prazos',
    'movimentacoes': 'movimentacoes',
    'movimentacoes web': 'movimentacoes',
    'andamentos': 'movimentacoes',
    'anotacoes': 'movimentacoes',
    'publicacoes': 'movimentacoes',
    'atendimentos': 'movimentacoes',
    'audiencias': 'movimentacoes',
    'documentos': 'documentos',
}

OPCOES_CAMPOS_IMPORTACAO = [
    ('', 'Ignorar coluna'),
    ('numero_processo', 'Numero CNJ'),
    ('titulo', 'Titulo'),
    ('descricao', 'Descricao'),
    ('tipo_acao', 'Tipo de acao'),
    ('tribunal', 'Tribunal'),
    ('vara', 'Vara'),
    ('comarca', 'Comarca'),
    ('valor_da_causa', 'Valor da causa'),
    ('status', 'Status'),
    ('fase', 'Fase'),
    ('data_criacao', 'Data de criacao'),
    ('parte_cliente', 'Parte cliente'),
    ('tipo_parte_cliente', 'Tipo da parte cliente'),
    ('contato_parte_cliente', 'Contato da parte cliente'),
    ('parte_adversa', 'Parte adversa'),
    ('tipo_parte_adversa', 'Tipo da parte adversa'),
    ('contato_parte_adversa', 'Contato da parte adversa'),
    ('outras_partes', 'Outras partes'),
    ('prazos', 'Prazos'),
    ('movimentacoes', 'Movimentacoes'),
    ('documentos', 'Documentos'),
]

COLUNAS_AGRUPAVEIS = {'prazos', 'movimentacoes', 'documentos', 'outras_partes'}
VALORES_VAZIOS_IMPORTACAO = {
    'sem parte cliente',
    'sem parte adversa',
    'sem parte',
    'nao informado',
    'não informado',
    'nao informada',
    'não informada',
}


def normalizar_cabecalho(valor):
    texto = str(valor or '').strip().lower()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(ch for ch in texto if unicodedata.category(ch) != 'Mn')
    texto = texto.replace('º', '').replace('°', '')
    for separador in ['-', '/', '.', '_', '\\', ':', ';', '(', ')', '[', ']']:
        texto = texto.replace(separador, ' ')
    return ' '.join(texto.split())


def identificar_coluna(coluna):
    cabecalho = normalizar_cabecalho(coluna)
    if not cabecalho:
        return None

    if 'cnj' in cabecalho and 'parte' not in cabecalho:
        return 'numero_processo'

    if 'processo' in cabecalho and 'sei' not in cabecalho:
        return 'numero_processo'

    return COLUNAS_IMPORTACAO.get(cabecalho)


def normalizar_valor_planilha(valor):
    if valor is None:
        return ''
    if hasattr(valor, 'strftime'):
        return valor.strftime('%Y-%m-%d')
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    if normalizar_cabecalho(texto) in {normalizar_cabecalho(v) for v in VALORES_VAZIOS_IMPORTACAO}:
        return ''
    return texto


def normalizar_valor_monetario(valor):
    texto = normalizar_valor_planilha(valor)
    if not texto:
        return ''

    texto = texto.replace('R$', '').replace(' ', '')
    if ',' in texto and '.' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    else:
        texto = texto.replace(',', '.')

    try:
        return str(Decimal(texto).quantize(Decimal('0.01')))
    except InvalidOperation:
        return normalizar_valor_planilha(valor)


def detectar_delimitador(texto):
    primeira_linha = texto.splitlines()[0] if texto.splitlines() else ''
    if primeira_linha.count(';') >= primeira_linha.count(','):
        return ';'
    return ','


def localizar_indice_cabecalho(linhas_brutas):
    indice_cabecalho = None
    for indice, linha in enumerate(linhas_brutas[:30]):
        if any(identificar_coluna(celula) == 'numero_processo' for celula in linha):
            indice_cabecalho = indice
            break

    return indice_cabecalho


def construir_linhas_com_cabecalho(linhas_brutas):
    indice_cabecalho = localizar_indice_cabecalho(linhas_brutas)
    if indice_cabecalho is None:
        return [], [], None

    cabecalho = [str(celula or '').strip() for celula in linhas_brutas[indice_cabecalho]]
    resultado = []
    for numero_linha, linha in enumerate(linhas_brutas[indice_cabecalho + 1:], start=indice_cabecalho + 2):
        resultado.append({
            '_numero_linha': numero_linha,
            **{
                cabecalho[indice]: valor
                for indice, valor in enumerate(linha)
                if indice < len(cabecalho)
            }
        })
    return resultado, list(cabecalho), indice_cabecalho


def ler_linhas_csv(arquivo):
    conteudo = arquivo.read()
    try:
        texto = conteudo.decode('utf-8-sig')
    except UnicodeDecodeError:
        texto = conteudo.decode('cp1252')

    leitor = csv.reader(io.StringIO(texto), delimiter=detectar_delimitador(texto))
    return construir_linhas_com_cabecalho(list(leitor))


def ler_linhas_xlsx(arquivo):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError('A importacao de .xlsx precisa da dependencia openpyxl instalada. Reinstale as dependencias ou envie CSV.') from exc

    arquivo.seek(0)
    workbook = load_workbook(arquivo, read_only=True, data_only=True)
    sheet = workbook.active
    return construir_linhas_com_cabecalho(list(sheet.iter_rows(values_only=True)))


def ler_planilha_processos(arquivo):
    nome = (arquivo.filename or '').lower()
    if nome.endswith('.csv'):
        linhas, _, _ = ler_linhas_csv(arquivo)
        return linhas
    if nome.endswith('.xlsx'):
        linhas, _, _ = ler_linhas_xlsx(arquivo)
        return linhas
    raise ValueError('Envie um arquivo .xlsx ou .csv.')


def ler_planilha_processos_detalhada(arquivo):
    nome = (arquivo.filename or '').lower()
    if nome.endswith('.csv'):
        linhas, cabecalho, indice_cabecalho = ler_linhas_csv(arquivo)
    elif nome.endswith('.xlsx'):
        linhas, cabecalho, indice_cabecalho = ler_linhas_xlsx(arquivo)
    else:
        raise ValueError('Envie um arquivo .xlsx ou .csv.')

    return {
        'linhas': linhas,
        'cabecalho': cabecalho,
        'indice_cabecalho': indice_cabecalho,
        'mapeamento_detectado': {
            coluna: identificar_coluna(coluna)
            for coluna in cabecalho
        },
    }


def mapear_linha(linha, mapeamento_colunas=None):
    dados = {}
    mapeamento_colunas = mapeamento_colunas or {}
    for coluna, valor in linha.items():
        if coluna == '_numero_linha':
            continue

        chave = mapeamento_colunas.get(coluna)
        if chave is None:
            chave = identificar_coluna(coluna)
        if not chave:
            continue

        if chave == 'valor_da_causa':
            valor_normalizado = normalizar_valor_monetario(valor)
        else:
            valor_normalizado = normalizar_valor_planilha(valor)

        if chave in COLUNAS_AGRUPAVEIS and dados.get(chave) and valor_normalizado:
            dados[chave] = f"{dados[chave]}\n{valor_normalizado}"
        elif valor_normalizado or chave not in dados:
            dados[chave] = valor_normalizado

    return dados


def completar_defaults_importacao(dados):
    numero = dados.get('numero_processo', '').strip()
    if numero:
        somente_digitos = ''.join(ch for ch in numero if ch.isdigit())
        if len(somente_digitos) == 20:
            dados['numero_processo'] = (
                f"{somente_digitos[:7]}-{somente_digitos[7:9]}."
                f"{somente_digitos[9:13]}.{somente_digitos[13]}."
                f"{somente_digitos[14:16]}.{somente_digitos[16:20]}"
            )

    if not dados.get('titulo') and dados.get('numero_processo'):
        dados['titulo'] = dados['numero_processo']
    if not dados.get('tipo_acao'):
        dados['tipo_acao'] = 'Nao informado'
    if not dados.get('status'):
        dados['status'] = 'ATIVO'

    return dados


def linha_sem_numero_deve_ser_ignorada(dados):
    return not (dados.get('numero_processo') or '').strip()


def preparar_processos_importacao(arquivo, mapeamento_colunas=None):
    linhas = ler_planilha_processos(arquivo)
    processos = []
    erros = []

    for indice_padrao, linha in enumerate(linhas, start=2):
        numero_linha = linha.get('_numero_linha', indice_padrao)
        dados = mapear_linha(linha, mapeamento_colunas=mapeamento_colunas)
        dados = completar_defaults_importacao(dados)
        if not any(dados.values()):
            continue
        if linha_sem_numero_deve_ser_ignorada(dados):
            continue

        try:
            processo = CadastroService.normalizar_processo_juridico(dados)
            partes, eventos = CadastroService.normalizar_vinculos_processo(dados)
            processos.append({
                'processo': processo,
                'partes': partes,
                'eventos': eventos,
                'linha': numero_linha,
            })
        except Exception as exc:
            erros.append(f"Linha {numero_linha}: {exc}")

    return processos, erros
